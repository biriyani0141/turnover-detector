"""
貸借・信用銘柄リスト変換スクリプト
実行方法:
  1. JPX公式からExcelをダウンロードし、プロジェクトルートに margin_list.xlsx として配置
  2. python scripts/build_margin_list.py

依存: openpyxl (pip install openpyxl)
出力: web/public/data/margin_list.json
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl が必要です。pip install openpyxl を実行してください。")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "margin_list.xlsx"
OUT_FILE = ROOT / "web" / "public" / "data" / "margin_list.json"

if not XLSX.exists():
    sys.exit(f"ERROR: {XLSX} が見つかりません。JPXからダウンロードして配置してください。")

wb = openpyxl.load_workbook(XLSX, data_only=True)

print("=== シート一覧 ===")
for name in wb.sheetnames:
    print(f"  シート: {name!r}")

print("\n=== 各シートの列構成（先頭3行） ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name!r} ---")
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > 3:
            break
        print(f"  行{row_idx}: {list(row)}")

# --- 列の自動検出 ---
# 銘柄コード列: 4〜5桁の数字が続く列を探す
# 区分列: "貸借" / "信用" / "制度信用" 等の文字列が入る列

def detect_columns(ws):
    """先頭20行を走査し、コード列・区分列のインデックスを返す"""
    code_col = None
    type_col = None

    # ヘッダー行を探す（最初の5行以内）
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        row_strs = [str(c) if c is not None else "" for c in row]
        if any("コード" in s or "code" in s.lower() for s in row_strs):
            header_row = row_idx
            break

    if header_row:
        row_vals = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        for i, v in enumerate(row_vals):
            s = str(v) if v is not None else ""
            if ("コード" in s or "code" in s.lower()) and code_col is None:
                code_col = i
            if ("区分" in s or "信用" in s or "貸借" in s or "種別" in s) and type_col is None:
                type_col = i
        if code_col is not None and type_col is not None:
            return code_col, type_col, header_row

    # ヘッダー行が見つからない場合はデータ行から推定
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        for i, v in enumerate(row):
            s = str(v).strip() if v is not None else ""
            if s.isdigit() and 4 <= len(s) <= 5 and code_col is None:
                code_col = i
            if s in ("貸借", "信用", "制度信用", "制度貸借") and type_col is None:
                type_col = i

    return code_col, type_col, None


stocks: dict[str, str] = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    code_col, type_col, header_row = detect_columns(ws)

    if code_col is None or type_col is None:
        print(f"\n[WARN] {sheet_name!r}: コード列/区分列を自動検出できませんでした。スキップします。")
        print(f"       code_col={code_col}, type_col={type_col}")
        continue

    start_row = (header_row + 1) if header_row else 1
    print(f"\n[INFO] {sheet_name!r}: code列={code_col}, type列={type_col}, データ開始行={start_row}")

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        code_val = row[code_col] if len(row) > code_col else None
        type_val = row[type_col] if len(row) > type_col else None

        if code_val is None or type_val is None:
            continue

        # コードは文字列のまま保持（数値変換・ゼロパディング変換は禁止）
        code_str = str(code_val).strip()
        type_str = str(type_val).strip()

        # 4〜5桁の数字または英数字コードのみ受け入れる
        if not (4 <= len(code_str) <= 5 and code_str.replace(".", "").isalnum()):
            continue
        if not type_str:
            continue

        stocks[code_str] = type_str

# バリデーション
count = len(stocks)
print(f"\n抽出件数: {count}")

if count < 1000:
    sys.exit(f"ERROR: 抽出件数が {count} 件で1000件未満です。列検出に失敗している可能性があります。処理を中断します。")

if OUT_FILE.exists():
    try:
        existing = json.loads(OUT_FILE.read_text(encoding="utf-8"))
        prev_count = len(existing.get("stocks", {}))
        if count < prev_count * 0.8:
            sys.exit(
                f"ERROR: 新件数 {count} が前回 {prev_count} の80%未満です。上書きせず中断します。"
            )
    except Exception:
        pass

import datetime
output = {
    "updated": datetime.date.today().isoformat(),
    "stocks": stocks,
}

OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
OUT_FILE.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"出力: {OUT_FILE} ({count}件)")
